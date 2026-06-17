from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.security import hash_id_number
from app.models.student import Student

EXPECTED_HEADERS = [
    "姓名",
    "学号",
    "身份证号",
    "班级",
    "宿舍",
    "辅导员",
    "辅导员电话",
    "班主任",
    "班主任电话",
    "代班",
    "代班电话",
    "代班班级",
]


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_header_map(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_cell(cell)
        if key:
            mapping[key] = idx
    return mapping


def _validate_headers(mapping: Dict[str, int]) -> List[str]:
    return [h for h in EXPECTED_HEADERS if h not in mapping]


def _cell(row_list: List[Any], hmap: Dict[str, int], name: str) -> str:
    idx = hmap[name]
    return _norm_cell(row_list[idx]) if idx < len(row_list) else ""


def _empty_to_none(s: str) -> str | None:
    return s if s else None


@dataclass
class ImportResult:
    imported: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)


def _skip_row(result: ImportResult, row_num: int, reason: str) -> None:
    result.skipped += 1
    result.errors.append({"row": row_num, "reason": reason})


def _build_student_payload(row_list: List[Any], hmap: Dict[str, int]) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": _cell(row_list, hmap, "姓名"),
        "id_number_hash": hash_id_number(_cell(row_list, hmap, "身份证号")),
        "class_name": _cell(row_list, hmap, "班级"),
        "dormitory": _empty_to_none(_cell(row_list, hmap, "宿舍")),
        "advisor_name": _empty_to_none(_cell(row_list, hmap, "辅导员")),
        "advisor_phone": _empty_to_none(_cell(row_list, hmap, "辅导员电话")),
        "class_teacher_name": _empty_to_none(_cell(row_list, hmap, "班主任")),
        "class_teacher_phone": _empty_to_none(_cell(row_list, hmap, "班主任电话")),
        "assistant_name": _empty_to_none(_cell(row_list, hmap, "代班")),
        "assistant_phone": _empty_to_none(_cell(row_list, hmap, "代班电话")),
        "assistant_class_name": _empty_to_none(_cell(row_list, hmap, "代班班级")),
        "role": "student",
    }
    if "照片" in hmap:
        payload["photo"] = _empty_to_none(_cell(row_list, hmap, "照片"))
    return payload


def _validate_data_row(
    row_list: List[Any],
    hmap: Dict[str, int],
    row_num: int,
    result: ImportResult,
) -> tuple[str, dict[str, Any]] | None:
    name = _cell(row_list, hmap, "姓名")
    sid = _cell(row_list, hmap, "学号")
    id_number = _cell(row_list, hmap, "身份证号")

    if not sid and not name and not id_number:
        _skip_row(result, row_num, "空行，已跳过")
        return None
    if not name:
        _skip_row(result, row_num, "姓名为空")
        return None
    if not sid:
        _skip_row(result, row_num, "学号为空")
        return None
    if not id_number:
        _skip_row(result, row_num, "身份证号为空")
        return None
    if not _cell(row_list, hmap, "班级"):
        _skip_row(result, row_num, "班级为空（必填）")
        return None

    return sid, _build_student_payload(row_list, hmap)


def _upsert_student_row(
    db: Session,
    sid: str,
    payload: dict[str, Any],
    result: ImportResult,
    row_num: int,
) -> None:
    existing = db.scalars(select(Student).where(Student.student_id == sid)).first()
    if existing and (existing.role or "").strip().lower() == "admin":
        _skip_row(result, row_num, "该行学号为系统管理员账号，已跳过")
        return
    if existing:
        for key, val in payload.items():
            setattr(existing, key, val)
        result.updated += 1
    else:
        db.add(Student(student_id=sid, **payload))
        result.imported += 1


def _process_import_row(
    db: Session,
    row_list: List[Any],
    hmap: Dict[str, int],
    row_num: int,
    result: ImportResult,
) -> None:
    parsed = _validate_data_row(row_list, hmap, row_num, result)
    if not parsed:
        return
    sid, payload = parsed
    _upsert_student_row(db, sid, payload, result, row_num)


def import_students_xlsx(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.errors.append({"row": 1, "reason": "文件为空"})
            result.skipped += 1
            return result

        hmap = _read_header_map(list(header_row))
        missing = _validate_headers(hmap)
        if missing:
            result.errors.append(
                {"row": 1, "reason": f"表头不完整，缺少：{', '.join(missing)}"}
            )
            result.skipped += 1
            return result

        row_num = 1
        for row in rows_iter:
            row_num += 1
            if row is None:
                continue
            _process_import_row(db, list(row), hmap, row_num, result)

        db.commit()
    finally:
        wb.close()

    return result
